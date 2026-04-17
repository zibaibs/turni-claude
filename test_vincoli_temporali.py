"""
Test per la gestione dei vincoli temporali in turnazione_completa.py

Copre:
  1. Parsing del foglio VincoliTemporali (datetime/date/str, ora come time/str)
  2. Retro-compat: foglio mancante o vuoto → constraints = {}
  3. Riga incompleta: Data fine vuota → end_date = start_date; Ora vuota → default 00:00-23:59
  4. shift_within_window via apply_time_constraints
  5. Vincoli multipli sovrapposti → intersezione
  6. Vincolo che svuota le opzioni → _forced_rest_by_constraints marca il giorno
  7. Operatore sconosciuto → warning + riga ignorata
"""
from __future__ import annotations

import sys
import tempfile
from datetime import datetime, date, time
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent))

from openpyxl import Workbook

from turnazione_completa import (
    Operator,
    TimeConstraint,
    apply_time_constraints,
    generate_shift_options,
    _forced_rest_by_constraints,
    load_input,
)

# ── helper: costruisci un workbook input minimo valido ────────────────────────

def build_minimal_workbook(vincoli_rows=None, include_vincoli_sheet=True):
    wb = Workbook()
    ws_p = wb.active
    ws_p.title = "Personale"
    ws_p.append(["Nome operatore", "Gruppo", "Ore giornaliere"])
    ws_p.append(["Mario Rossi", "Zetema", 8])
    ws_p.append(["Anna Bianchi", "Zetema", 6])
    ws_p.append(["Luigi Verdi", "Zetema", 4])
    ws_p.append(["Carla Neri", "Non Zetema", 8])

    ws_per = wb.create_sheet("Periodo")
    ws_per.append(["Data inizio", "Data fine"])
    # Lun 5/5/2025 - Dom 11/5/2025 (1 settimana)
    ws_per.append([datetime(2025, 5, 5), datetime(2025, 5, 11)])

    ws_c = wb.create_sheet("Copertura")
    ws_c.append(["Fascia", "Lun", "Mar", "Mer", "Gio", "Ven", "Sab", "Dom"])
    for h in range(7, 20):
        ws_c.append([f"{h:02d}:00-{h+1:02d}:00", 1, 1, 1, 1, 1, 1, 1])

    ws_a = wb.create_sheet("Assenze")
    ws_a.append(["Nome operatore", "Data assenza", "Tipo assenza"])

    if include_vincoli_sheet:
        ws_v = wb.create_sheet("VincoliTemporali")
        ws_v.append(["Nome operatore", "Data inizio", "Data fine", "Ora inizio", "Ora fine", "Note"])
        for row in vincoli_rows or []:
            ws_v.append(row)
    return wb


def save_and_load(wb):
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = Path(f.name)
    wb.save(path)
    try:
        return load_input(path)
    finally:
        path.unlink(missing_ok=True)


# ─── Tests ───────────────────────────────────────────────────────────────────

def test_parsing_base():
    wb = build_minimal_workbook(vincoli_rows=[
        ["Mario Rossi", datetime(2025, 5, 5), datetime(2025, 5, 11), "07:00", "14:00", "mattina"],
        ["Anna Bianchi", datetime(2025, 5, 7), datetime(2025, 5, 7), "13:00", "20:00", ""],
    ])
    _, _, _, _, constraints = save_and_load(wb)
    assert "Mario Rossi" in constraints
    assert len(constraints["Mario Rossi"]) == 1
    c = constraints["Mario Rossi"][0]
    assert c.start_date == datetime(2025, 5, 5)
    assert c.end_date == datetime(2025, 5, 11)
    assert c.window_start_min == 7 * 60
    assert c.window_end_min == 14 * 60
    assert c.note == "mattina"
    assert constraints["Anna Bianchi"][0].window_start_min == 13 * 60
    print("  ✓ parsing base OK")


def test_sheet_mancante():
    wb = build_minimal_workbook(include_vincoli_sheet=False)
    _, _, _, _, constraints = save_and_load(wb)
    assert constraints == {}
    print("  ✓ foglio mancante → constraints vuoti (retro-compat)")


def test_riga_incompleta():
    wb = build_minimal_workbook(vincoli_rows=[
        ["Mario Rossi", datetime(2025, 5, 6), None, None, None, "solo nota"],
    ])
    _, _, _, _, constraints = save_and_load(wb)
    c = constraints["Mario Rossi"][0]
    assert c.start_date == c.end_date == datetime(2025, 5, 6)
    assert c.window_start_min == 0
    assert c.window_end_min == 24 * 60 - 1
    print("  ✓ Data fine/ore vuote → default applicati")


def test_apply_time_constraints_basic():
    op = Operator("Test", "Zetema", 4)
    opts = generate_shift_options(op, 0)  # lun 07-11 → 16-20 (daily_hours=4 range 7..17)
    # Vincolo 07:00-14:00 → restano solo shift che finiscono ≤14:00
    cs = {"Test": [TimeConstraint(datetime(2025, 5, 5), datetime(2025, 5, 11), 7*60, 14*60)]}
    filtered = apply_time_constraints(opts, "Test", datetime(2025, 5, 5), cs)
    assert filtered, "filtro troppo aggressivo"
    for sh in filtered:
        start, end = sh.split("-")
        assert start <= "14:00" and end <= "14:00"
    # Giorno fuori periodo → opzioni invariate
    out = apply_time_constraints(opts, "Test", datetime(2025, 4, 1), cs)
    assert out == opts
    print("  ✓ apply_time_constraints filtra finestra e rispetta periodo")


def test_vincoli_multipli_intersezione():
    op = Operator("Test", "Zetema", 4)
    opts = generate_shift_options(op, 0)
    cs = {"Test": [
        TimeConstraint(datetime(2025, 5, 5), datetime(2025, 5, 11), 7*60, 14*60),
        TimeConstraint(datetime(2025, 5, 5), datetime(2025, 5, 11), 9*60, 17*60),
    ]}
    filtered = apply_time_constraints(opts, "Test", datetime(2025, 5, 6), cs)
    # Intersezione: 09:00-14:00 → shift di 4h validi start in [9..10]
    for sh in filtered:
        start, end = sh.split("-")
        assert start >= "09:00" and end <= "14:00"
    print(f"  ✓ intersezione vincoli sovrapposti ({len(filtered)} opzioni valide)")


def test_constraint_empties_options_forces_rest():
    # Op 8h con vincolo 09-13 → impossibile (nessuno shift 8h entra in 4h)
    op = Operator("Test8h", "Zetema", 8)
    week_days = [datetime(2025, 5, 5) + __import__("datetime").timedelta(days=i) for i in range(7)]
    cs = {"Test8h": [TimeConstraint(datetime(2025, 5, 5), datetime(2025, 5, 11), 9*60, 13*60)]}
    forced = _forced_rest_by_constraints([op], week_days, cs)
    # Lun-sab devono essere forzati (dom il daily=8 non ha opzioni di default → skip)
    assert 0 in forced["Test8h"] and 5 in forced["Test8h"]
    print(f"  ✓ vincolo incompatibile forza rest giorni {sorted(forced['Test8h'])}")


def test_operatore_sconosciuto_ignorato():
    wb = build_minimal_workbook(vincoli_rows=[
        ["Fantasma", datetime(2025, 5, 5), datetime(2025, 5, 11), "07:00", "14:00", ""],
        ["Mario Rossi", datetime(2025, 5, 5), datetime(2025, 5, 11), "07:00", "14:00", ""],
    ])
    _, _, _, _, constraints = save_and_load(wb)
    assert "Fantasma" not in constraints
    assert "Mario Rossi" in constraints
    print("  ✓ operatore sconosciuto ignorato con warning")


def test_end_before_start_ignored():
    wb = build_minimal_workbook(vincoli_rows=[
        ["Mario Rossi", datetime(2025, 5, 11), datetime(2025, 5, 5), "07:00", "14:00", ""],
    ])
    _, _, _, _, constraints = save_and_load(wb)
    assert constraints == {} or "Mario Rossi" not in constraints
    print("  ✓ Data fine<Data inizio → vincolo scartato")


def test_ore_invalid_ignored():
    wb = build_minimal_workbook(vincoli_rows=[
        ["Mario Rossi", datetime(2025, 5, 5), datetime(2025, 5, 11), "14:00", "07:00", ""],
    ])
    _, _, _, _, constraints = save_and_load(wb)
    assert "Mario Rossi" not in constraints
    print("  ✓ Ora fine<=Ora inizio → vincolo scartato")


# ─── Runner ──────────────────────────────────────────────────────────────────

def run():
    tests = [
        test_parsing_base,
        test_sheet_mancante,
        test_riga_incompleta,
        test_apply_time_constraints_basic,
        test_vincoli_multipli_intersezione,
        test_constraint_empties_options_forces_rest,
        test_operatore_sconosciuto_ignorato,
        test_end_before_start_ignored,
        test_ore_invalid_ignored,
    ]
    failed = 0
    for t in tests:
        try:
            print(f"[RUN] {t.__name__}")
            t()
        except AssertionError as e:
            print(f"  ✗ FAIL: {e}")
            failed += 1
        except Exception as e:
            import traceback
            print(f"  ✗ ERROR: {e}")
            traceback.print_exc()
            failed += 1
    print()
    if failed:
        print(f"✗ {failed} test falliti")
        sys.exit(1)
    print(f"✓ Tutti i {len(tests)} test passati")


if __name__ == "__main__":
    run()
