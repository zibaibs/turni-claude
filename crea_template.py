"""
Genera il file input_turni_template.xlsx con la struttura esatta
richiesta da turnazione_completa.py.
"""
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True)
EXAMPLE_FILL = PatternFill("solid", fgColor="D9E1F2")
NOTE_FONT = Font(italic=True, color="7F7F7F")
THIN = Side(style="thin")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def header(ws, row, col, text, width=None):
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = BORDER
    if width and hasattr(ws, "column_dimensions"):
        ws.column_dimensions[get_column_letter(col)].width = width


def example(ws, row, col, value, note=False):
    cell = ws.cell(row=row, column=col, value=value)
    cell.fill = EXAMPLE_FILL
    cell.border = BORDER
    cell.alignment = Alignment(horizontal="center", vertical="center")
    if note:
        cell.font = NOTE_FONT
    return cell


wb = Workbook()

# ── 1. Personale ─────────────────────────────────────────────────────────────
ws_p = wb.active
ws_p.title = "Personale"
ws_p.row_dimensions[1].height = 30

headers_p = ["Nome operatore", "Gruppo", "Ore giornaliere"]
widths_p = [25, 18, 18]
for c, (h, w) in enumerate(zip(headers_p, widths_p), start=1):
    header(ws_p, 1, c, h, w)

# Note row (row 2) – shown as example placeholder
examples_p = [
    ("Mario Rossi", "Zetema", 8),
    ("Anna Bianchi", "Zetema", 6),
    ("Luigi Verdi", "Non Zetema", 8),
    ("Carla Neri", "Non Zetema", 4),
]
for r, (name, group, ore) in enumerate(examples_p, start=2):
    example(ws_p, r, 1, name)
    cell_g = example(ws_p, r, 2, group)
    cell_o = example(ws_p, r, 3, ore)

# Note below table
note_row = len(examples_p) + 3
ws_p.cell(row=note_row, column=1,
          value="⚠ Gruppo: solo 'Zetema' o 'Non Zetema'  |  Ore: 4, 6 oppure 8").font = NOTE_FONT
ws_p.cell(row=note_row + 1, column=1,
          value="Eliminare le righe di esempio e inserire i dati reali.").font = NOTE_FONT

# ── 2. Periodo ───────────────────────────────────────────────────────────────
ws_per = wb.create_sheet("Periodo")
ws_per.row_dimensions[1].height = 30

headers_per = ["Data inizio (lunedì)", "Data fine (domenica)"]
widths_per = [24, 24]
for c, (h, w) in enumerate(zip(headers_per, widths_per), start=1):
    header(ws_per, 1, c, h, w)

# Example: first week of a month
start_ex = datetime(2025, 4, 7)   # lunedì
end_ex = datetime(2025, 4, 27)    # domenica (3 settimane)
c_start = example(ws_per, 2, 1, start_ex)
c_start.number_format = "DD/MM/YYYY"
c_end = example(ws_per, 2, 2, end_ex)
c_end.number_format = "DD/MM/YYYY"

note_per = ws_per.cell(row=4, column=1,
                        value="⚠ La data di inizio deve essere un LUNEDÌ."
                              "  Il periodo deve coprire settimane intere (multiplo di 7 giorni).")
note_per.font = NOTE_FONT

# ── 3. Copertura ─────────────────────────────────────────────────────────────
ws_c = wb.create_sheet("Copertura")
ws_c.row_dimensions[1].height = 30

day_headers = ["Fascia oraria", "Lun", "Mar", "Mer", "Gio", "Ven", "Sab", "Dom"]
day_widths = [20, 8, 8, 8, 8, 8, 8, 8]
for c, (h, w) in enumerate(zip(day_headers, day_widths), start=1):
    header(ws_c, 1, c, h, w)

# Optional block-start marker (datetime in col A)
block_marker = ws_c.cell(row=2, column=1, value=datetime(2025, 4, 7))
block_marker.number_format = "DD/MM/YYYY"
block_marker.fill = PatternFill("solid", fgColor="FFF2CC")
block_marker.font = Font(bold=True)
block_marker.border = BORDER
block_marker.alignment = Alignment(horizontal="center", vertical="center")
ws_c.cell(row=2, column=2,
          value="← Riga opzionale: data (lunedì) da cui vale questo fabbisogno").font = NOTE_FONT

# Demand rows for hours 7–19
demand_examples = {
    7:  [1, 1, 1, 1, 1, 0, 0],
    8:  [2, 2, 2, 2, 2, 1, 0],
    9:  [3, 3, 3, 3, 3, 2, 1],
    10: [3, 3, 3, 3, 3, 2, 1],
    11: [3, 3, 3, 3, 3, 2, 1],
    12: [2, 2, 2, 2, 2, 1, 0],
    13: [1, 1, 1, 1, 1, 0, 0],
    14: [1, 1, 1, 1, 1, 0, 0],
    15: [2, 2, 2, 2, 2, 1, 0],
    16: [2, 2, 2, 2, 2, 1, 0],
    17: [2, 2, 2, 2, 2, 1, 0],
    18: [1, 1, 1, 1, 1, 0, 0],
    19: [1, 1, 1, 1, 1, 0, 0],
}
for i, (hour, vals) in enumerate(demand_examples.items()):
    r = i + 3  # righe 3-15
    slot_label = f"{hour:02d}:00-{hour+1:02d}:00"
    example(ws_c, r, 1, slot_label)
    for j, v in enumerate(vals):
        example(ws_c, r, j + 2, v)

note_c_row = len(demand_examples) + 4
ws_c.cell(row=note_c_row, column=1,
          value="⚠ Valori = numero minimo di operatori presenti in quell'ora per quel giorno.").font = NOTE_FONT
ws_c.cell(row=note_c_row + 1, column=1,
          value="Per più blocchi di fabbisogno (date diverse), inserire una nuova riga-data e ripetere la tabella.").font = NOTE_FONT

# ── 4. Assenze ───────────────────────────────────────────────────────────────
ws_a = wb.create_sheet("Assenze")
ws_a.row_dimensions[1].height = 30

headers_a = ["Nome operatore", "Data assenza", "Tipo assenza"]
widths_a = [25, 20, 18]
for c, (h, w) in enumerate(zip(headers_a, widths_a), start=1):
    header(ws_a, 1, c, h, w)

# Example rows
absences_ex = [
    ("Mario Rossi", datetime(2025, 4, 9), "Ferie"),
    ("Anna Bianchi", datetime(2025, 4, 14), "Malattia"),
]
for r, (name, date, tipo) in enumerate(absences_ex, start=2):
    example(ws_a, r, 1, name)
    cell_d = example(ws_a, r, 2, date)
    cell_d.number_format = "DD/MM/YYYY"
    example(ws_a, r, 3, tipo)

note_a = ws_a.cell(row=5, column=1,
                    value="⚠ Inserire una riga per ogni giorno di assenza."
                          "  Il nome deve corrispondere esattamente al foglio Personale."
                          "  Tipo: es. Ferie, Malattia, Permesso (opzionale, default: Assenza).")
note_a.font = NOTE_FONT

# ── Salva ─────────────────────────────────────────────────────────────────────
out = "input_turni_template.xlsx"
wb.save(out)
print(f"Template salvato: {out}")
