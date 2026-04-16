from __future__ import annotations

import random
import sys
from dataclasses import dataclass
from datetime import date, datetime, time, timedelta
from pathlib import Path
import itertools
import re
from typing import Dict, List, Set, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.formatting.rule import CellIsRule, FormulaRule

# PyInstaller: usa CWD (cartella da cui viene lanciato l'exe/bat)
# così input_turni.xlsx e output/ si trovano sempre accanto al bat/exe.
if getattr(sys, "frozen", False):
    _HERE = Path.cwd()
else:
    _HERE = Path(__file__).parent
INPUT_FILE = _HERE / "input_turni.xlsx"
OUTPUT_FILE = _HERE / "output" / "turnazione_generata.xlsx"
DAY_LABELS = ["Lun", "Mar", "Mer", "Gio", "Ven", "Sab", "Dom"]
SHIFT_RE = re.compile(r"^(\d{2}:\d{2})-(\d{2}:\d{2})(?:\n\(pausa (\d{2}:\d{2})\))?$")

# Weight applied to work-history counts when scoring rest-day rotation.
# Higher value = stronger rotation pressure vs. demand balance.
ROTATION_WEIGHT = 50

# Retry loop: rigenera schedule se i delta negativi sono eccessivi.
MAX_SCHEDULE_RETRIES = 10      # tentativi massimi per settimana
DEFICIT_TOLERANCE = -2         # deficit giornaliero accettabile (persona-ore)


@dataclass
class Operator:
    name: str
    group: str
    daily_hours: int


def parse_hhmm(value: str) -> time:
    return datetime.strptime(value, "%H:%M").time()


def to_minutes(value: time) -> int:
    return value.hour * 60 + value.minute


def parse_shift(shift: str) -> Tuple[int, int, int | None]:
    match = SHIFT_RE.match(shift)
    if not match:
        raise ValueError(f"Formato turno non valido: '{shift}'")
    start = to_minutes(parse_hhmm(match.group(1)))
    end = to_minutes(parse_hhmm(match.group(2)))
    pause_start = to_minutes(parse_hhmm(match.group(3))) if match.group(3) else None
    return start, end, pause_start


def shift_to_hours(shift: str) -> Set[int]:
    start, end, pause = parse_shift(shift)
    covered = set(range(start // 60, end // 60))
    if pause is not None:
        covered.discard(pause // 60)
    return covered


def format_shift(start_h: int, end_h: int, pause_h: int | None) -> str:
    if pause_h is None:
        return f"{start_h:02d}:00-{end_h:02d}:00"
    return f"{start_h:02d}:00-{end_h:02d}:00\n(pausa {pause_h:02d}:00)"


def is_working_shift(s: str) -> bool:
    """True if s is a valid shift string (not RIP, Ferie, Malattia, etc.)."""
    return bool(SHIFT_RE.match(s))


def parse_coverage_slot(slot_text: str) -> int | None:
    match = re.search(r"(\d{1,2})\D+(\d{1,2})", slot_text)
    if not match:
        return None
    return int(match.group(1))


def load_input(
    path: Path,
) -> Tuple[List[Operator], List[List[datetime]], List[Tuple[datetime, Dict[int, List[int]]]], Dict[str, Dict[datetime, str]]]:
    """Load and validate the input workbook.

    Returns:
        operators       – list of Operator objects
        weeks           – list of weeks; each week is a list of 7 datetime objects (Mon–Sun)
        demand_schedule – list of (start_date, demand_matrix) sorted ascending;
                          each matrix covers all weeks from start_date until the next entry
        absences        – {operator_name: {date: tipo_label}} e.g. {"Mario": {datetime(...): "Ferie"}}
    """
    if not path.exists():
        raise FileNotFoundError(f"File input non trovato: {path}")

    wb = load_workbook(path)
    for required in ["Personale", "Copertura", "Periodo", "Assenze"]:
        if required not in wb.sheetnames:
            raise ValueError(f"Manca il foglio obbligatorio '{required}' nel file input.")

    # --- Personale ---
    ws_personale = wb["Personale"]
    operators: List[Operator] = []
    for row in ws_personale.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        name, group, hours = row[:3]
        if group not in {"Zetema", "Non Zetema"}:
            raise ValueError(f"Gruppo non valido per '{name}': {group}")
        if hours not in {4, 6, 8}:
            raise ValueError(f"Ore giornaliere non valide per '{name}': {hours}")
        operators.append(Operator(str(name).strip(), str(group).strip(), int(hours)))

    # --- Periodo (multi-week) ---
    ws_periodo = wb["Periodo"]
    start_date = ws_periodo["A2"].value
    end_date = ws_periodo["B2"].value
    if not isinstance(start_date, datetime) or not isinstance(end_date, datetime):
        raise ValueError("Date non valide nel foglio Periodo (celle A2/B2).")
    total_days = (end_date - start_date).days + 1
    if total_days % 7 != 0:
        raise ValueError("Il periodo deve essere un multiplo di 7 giorni (lun–dom).")
    if start_date.weekday() != 0:
        raise ValueError("La data di inizio deve essere un lunedì.")

    base = start_date.replace(hour=0, minute=0, second=0, microsecond=0)
    num_weeks = total_days // 7
    weeks = [
        [base + timedelta(days=w * 7 + i) for i in range(7)]
        for w in range(num_weeks)
    ]

    # --- Copertura (one or more demand matrices, each valid from a start date) ---
    # Format: an optional row with a datetime in column A marks the start of a new block.
    # If no datetime markers are present, the entire table is treated as a single matrix
    # valid for all weeks.
    ws_copertura = wb["Copertura"]

    def _empty_demand() -> Dict[int, List[int]]:
        return {hour: [0] * 7 for hour in range(7, 20)}

    demand_schedule: List[Tuple[datetime, Dict[int, List[int]]]] = []
    current_date: datetime | None = None
    current_demand: Dict[int, List[int]] | None = None

    for row in ws_copertura.iter_rows(min_row=2, max_col=8, values_only=True):
        first = row[0]
        if first is None:
            continue
        if isinstance(first, datetime):
            # Save the previous block (if any) and start a new one
            if current_date is not None and current_demand is not None:
                demand_schedule.append((current_date, current_demand))
            current_date = first.replace(hour=0, minute=0, second=0, microsecond=0)
            current_demand = _empty_demand()
        else:
            # Demand data row: add to the current block (or bootstrap an implicit block)
            if current_demand is None:
                current_date = datetime.min
                current_demand = _empty_demand()
            start_hour = parse_coverage_slot(str(first))
            if start_hour is None or start_hour not in current_demand:
                continue
            current_demand[start_hour] = [int(v or 0) for v in row[1:8]]

    if current_date is not None and current_demand is not None:
        demand_schedule.append((current_date, current_demand))

    if not demand_schedule:
        raise ValueError("Nessun fabbisogno trovato nel foglio Copertura.")

    demand_schedule.sort(key=lambda x: x[0])

    # --- Assenze (stored as absolute dates for multi-week lookup) ---
    # Structure: {operator_name: {date: tipo_label}}
    absences: Dict[str, Dict[datetime, str]] = {}
    known_op_names = {op.name for op in operators}
    ws_assenze = wb["Assenze"]
    for row_idx, row in enumerate(ws_assenze.iter_rows(min_row=2, values_only=True), start=2):
        if not row[0] or not row[1]:
            continue
        name = str(row[0]).strip()
        raw_date = row[1]
        tipo = str(row[2]).strip() if len(row) > 2 and row[2] else "Assenza"

        # Normalizzazione tipo data: supporta datetime, date e stringhe comuni
        if isinstance(raw_date, datetime):
            absence_date = raw_date.replace(hour=0, minute=0, second=0, microsecond=0)
        elif isinstance(raw_date, date):
            absence_date = datetime(raw_date.year, raw_date.month, raw_date.day)
        elif isinstance(raw_date, str):
            for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%d.%m.%Y"):
                try:
                    absence_date = datetime.strptime(raw_date.strip(), fmt)
                    break
                except ValueError:
                    continue
            else:
                print(
                    f"[AVVISO] Assenze riga {row_idx}: data non riconosciuta per '{name}'"
                    f": '{raw_date}' — assenza ignorata."
                )
                continue
        else:
            print(
                f"[AVVISO] Assenze riga {row_idx}: tipo data non supportato per '{name}'"
                f": {type(raw_date)} — assenza ignorata."
            )
            continue

        if name not in known_op_names:
            print(
                f"[AVVISO] Assenze riga {row_idx}: operatore '{name}' non trovato nel foglio"
                f" Personale — assenza ignorata."
            )
            continue

        absences.setdefault(name, {})[absence_date] = tipo

    return operators, weeks, demand_schedule, absences


def get_demand_for_week(
    demand_schedule: List[Tuple[datetime, Dict[int, List[int]]]],
    week_start: datetime,
) -> Dict[int, List[int]]:
    """Return the demand matrix valid for the given week start date.

    Uses the entry whose start_date is the latest one still <= week_start.
    """
    result = demand_schedule[0][1]
    for start_date, matrix in demand_schedule:
        if start_date <= week_start:
            result = matrix
        else:
            break
    return result


def get_week_absences(
    absences: Dict[str, Dict[datetime, str]],
    week_days: List[datetime],
) -> Dict[str, Dict[int, str]]:
    """Convert date-keyed absences to day-index (0–6) absences for a given week.

    Returns {operator_name: {day_idx: tipo_label}}.
    """
    week_map = {
        d.replace(hour=0, minute=0, second=0, microsecond=0): i
        for i, d in enumerate(week_days)
    }
    result: Dict[str, Dict[int, str]] = {}
    for name, date_map in absences.items():
        for d, tipo in date_map.items():
            idx = week_map.get(d)
            if idx is not None:
                result.setdefault(name, {})[idx] = tipo
    return result


def pick_sunday_workers(
    zetema_ops: List[Operator],
    absences: Dict[str, Dict[int, str]],
    sunday_history: Dict[str, int],
) -> Set[str]:
    """Select which Zetema operators work on Sunday.

    Uses rotation: among all valid combinations (6h+4h or 4h+4h+4h), those whose
    members have collectively worked the fewest previous Sundays are preferred.
    Ties are broken randomly so different combinations share the load over time.
    """
    available = [op for op in zetema_ops if 6 not in absences.get(op.name, {})]
    six_hour = [op for op in available if op.daily_hours == 6]
    four_hour = [op for op in available if op.daily_hours == 4]

    candidates: List[Tuple[int, Set[str]]] = []

    for s in six_hour:
        for f in four_hour:
            if s.name != f.name:
                score = sunday_history.get(s.name, 0) + sunday_history.get(f.name, 0)
                candidates.append((score, {s.name, f.name}))

    for combo in itertools.combinations(four_hour, 3):
        score = sum(sunday_history.get(op.name, 0) for op in combo)
        candidates.append((score, {op.name for op in combo}))

    if not candidates:
        raise ValueError("Impossibile coprire la domenica con combinazioni consentite (6+4 o 4+4+4).")

    min_score = min(s for s, _ in candidates)
    best = [combo for s, combo in candidates if s == min_score]
    return random.choice(best)


def generate_rest_days(
    operators: List[Operator],
    demand_by_hour: Dict[int, List[int]],
    absences: Dict[str, Dict[int, str]],
    sunday_workers: Set[str],
    work_history: Dict[str, Dict[int, int]],
) -> Dict[str, Set[int]]:
    """Assign 2 rest days per operator for the week.

    Rotation logic:
    - Days an operator has worked frequently in previous weeks get a higher
      rotation bonus, making them more attractive as rest days this week.
    - For Zetema, among all valid combinations (≥10h coverage each lun–sab day),
      the best-scoring ones are collected and one is picked at random, ensuring
      variety across runs.
    - For Non-Zetema, Saturday rotation is naturally rewarded via the same
      history-based scoring.
    """
    rest_days: Dict[str, Set[int]] = {}
    weekday_demand = [sum(demand_by_hour[h][d] for h in range(7, 20)) for d in range(6)]

    zetema_ops = [op for op in operators if op.group == "Zetema"]
    non_zetema_ops = [op for op in operators if op.group == "Non Zetema"]

    # --- Zetema rest days ---
    weekday_options: Dict[str, List[Tuple[int, ...]]] = {}
    for op in zetema_ops:
        forced = set(d for d in absences.get(op.name, {}) if d < 6)
        if op.name in sunday_workers:
            if 6 in absences.get(op.name, {}):
                raise ValueError(f"'{op.name}' e' assegnato domenica ma risulta assente.")
            required_weekday_rip = 2
        else:
            required_weekday_rip = 1
            rest_days[op.name] = {6}

        need_extra = required_weekday_rip - len(forced)
        if need_extra <= 0:
            # Assenze coprono (o superano) i riposi richiesti: nessun giorno extra da cercare
            weekday_options[op.name] = [tuple(sorted(forced))]
        else:
            available = [d for d in range(6) if d not in forced]
            weekday_options[op.name] = []
            for combo in itertools.combinations(available, need_extra):
                combined = tuple(sorted(set(combo) | forced))
                if len(combined) >= required_weekday_rip:
                    weekday_options[op.name].append(combined)
            if not weekday_options[op.name]:
                raise ValueError(f"Nessuna combinazione RIP valida per '{op.name}'.")

    zetema_names = [op.name for op in zetema_ops]

    # Cap the search space to prevent exponential slowdown with many Zetema operators.
    # Below the threshold do an exhaustive shuffled search; above it, random-sample.
    _MAX_RIP_CANDIDATES = 20_000
    _rip_product_size = 1
    for _n in zetema_names:
        _rip_product_size *= len(weekday_options[_n])

    if _rip_product_size <= _MAX_RIP_CANDIDATES:
        rip_candidates = list(itertools.product(*(weekday_options[name] for name in zetema_names)))
        random.shuffle(rip_candidates)
    else:
        rip_candidates = [
            tuple(random.choice(weekday_options[name]) for name in zetema_names)
            for _ in range(_MAX_RIP_CANDIDATES)
        ]

    best_score: float | None = None
    best_candidate: Dict[str, Set[int]] | None = None

    for candidate in rip_candidates:
        candidate_map = {name: set(days) for name, days in zip(zetema_names, candidate)}

        day_hours = [0] * 6
        for op in zetema_ops:
            for day_idx in range(6):
                if day_idx not in candidate_map[op.name]:
                    day_hours[day_idx] += op.daily_hours

        # Penalità per copertura insufficiente (non scarta: permette il meglio possibile con assenze forzate)
        coverage_penalty = sum(max(0, 10 - h) for h in day_hours) * 100_000

        # Primary: minimise resting on high-demand days (scaled up)
        demand_score = sum(
            weekday_demand[d]
            for name in zetema_names
            for d in candidate_map[name]
        )
        # Rotation bonus: reward resting on days the operator has worked most
        rotation_bonus = sum(
            work_history.get(name, {}).get(d, 0)
            for name in zetema_names
            for d in candidate_map[name]
        ) * ROTATION_WEIGHT

        score = demand_score * 1000 - rotation_bonus + coverage_penalty  # lower is better

        if best_score is None or score < best_score:
            best_score = score
            best_candidate = candidate_map

    if best_candidate is None:
        raise ValueError("Impossibile assegnare RIP Zetema con copertura 09:00-19:00 lun-sab.")

    for op in zetema_ops:
        rest_days.setdefault(op.name, set()).update(best_candidate[op.name])

    # --- Non-Zetema rest days ---
    nonz_rest_count = [0] * 6
    for op in non_zetema_ops:
        forced = set(absences.get(op.name, {}))  # keys are day indices
        forced.add(6)  # Sunday is always rest

        # Se assenze > 2, l'operatore lavora meno di 5 giorni (gestito senza crash)
        history = work_history.get(op.name, {})
        while len(forced) < 2:
            cands = [d for d in range(6) if d not in forced]

            def rest_score(d: int, _h: Dict[int, int] = history) -> float:
                # Prefer: less-used rest slots + low demand + days worked frequently
                return (
                    nonz_rest_count[d] * 1000
                    + weekday_demand[d]
                    - _h.get(d, 0) * ROTATION_WEIGHT
                )

            candidate = min(cands, key=rest_score) if cands else None
            if candidate is None:
                raise ValueError(f"Impossibile completare i riposi per '{op.name}'.")
            forced.add(candidate)
            nonz_rest_count[candidate] += 1
        rest_days[op.name] = forced

    # Final check: enough total hours available every lun–sab
    for day_idx in range(6):
        available_hours = sum(
            op.daily_hours for op in operators if day_idx not in rest_days[op.name]
        )
        if available_hours < 13:
            print(
                f"[AVVISO] Ore disponibili insufficienti in {DAY_LABELS[day_idx]} "
                f"({available_hours}h) — assenze non assorbibili, copertura ridotta."
            )
    return rest_days


def generate_shift_options(op: Operator, day_idx: int) -> List[str]:
    if day_idx == 6:
        if op.daily_hours == 6:
            return [format_shift(start, start + 6, None) for start in range(9, 14)]
        if op.daily_hours == 4:
            return [format_shift(start, start + 4, None) for start in range(9, 16)]
        return []

    if op.daily_hours == 4:
        return [format_shift(start, start + 4, None) for start in range(7, 17)]
    if op.daily_hours == 6:
        return [format_shift(start, start + 6, None) for start in range(7, 15)]

    options: List[str] = []
    for start in range(7, 12):
        for pause_after in (3, 4, 5):
            pause_hour = start + pause_after
            options.append(format_shift(start, start + 9, pause_hour))
    return options


def build_schedule(
    operators: List[Operator],
    demand_by_hour: Dict[int, List[int]],
    rest_days: Dict[str, Set[int]],
    sunday_workers: Set[str],
    prev_preferred: Dict[str, str] | None = None,
) -> Dict[str, List[str]]:
    """Assign concrete shifts to every working day for every operator.

    prev_preferred – dominant shifts from the previous week; used as a strong
    anchor to keep each operator on the same recurring shift across weeks.
    """
    schedule: Dict[str, List[str]] = {op.name: ["RIP"] * 7 for op in operators}
    # Pre-populate from last week so the first day of each week already has a
    # preferred shift and doesn't drift purely on demand.
    preferred_shift: Dict[str, str] = dict(prev_preferred) if prev_preferred else {}

    for day_idx in range(7):
        coverage = {h: 0 for h in range(7, 20)}
        z_coverage = {h: 0 for h in range(9, 19)}
        workers = [op for op in operators if day_idx not in rest_days[op.name]]
        workers.sort(key=lambda op: (op.group != "Zetema", -op.daily_hours))

        if day_idx == 6:
            workers = [op for op in workers if op.name in sunday_workers]

        zetema_workers = [op for op in workers if op.group == "Zetema"]
        if zetema_workers:
            option_buckets = [generate_shift_options(op, day_idx) for op in zetema_workers]
            best_combo: Tuple[str, ...] | None = None
            best_combo_score = -10**9

            # Cap the search space to prevent exponential slowdown with many Zetema operators.
            _MAX_SHIFT_COMBOS = 5_000
            _shift_combo_size = 1
            for _opts in option_buckets:
                _shift_combo_size *= len(_opts)

            if _shift_combo_size <= _MAX_SHIFT_COMBOS:
                shift_candidates: List[Tuple[str, ...]] = list(itertools.product(*option_buckets))
                random.shuffle(shift_candidates)
            else:
                shift_candidates = [
                    tuple(random.choice(opts) for opts in option_buckets)
                    for _ in range(_MAX_SHIFT_COMBOS)
                ]

            for combo in shift_candidates:
                test_cov = {h: 0 for h in range(9, 19)}
                combo_score = 0
                for op, shift in zip(zetema_workers, combo):
                    covered = shift_to_hours(shift)
                    for h in covered:
                        if 9 <= h < 19:
                            test_cov[h] += 1
                        unmet = max(0, demand_by_hour.get(h, [0] * 7)[day_idx] - coverage[h])
                        combo_score += unmet * 8 + 1
                    # Strong consistency bonus: same shift as the running preferred
                    if preferred_shift.get(op.name) == shift:
                        combo_score += 300
                    # Cross-week consistency: same shift as last week's dominant
                    if prev_preferred and prev_preferred.get(op.name) == shift:
                        combo_score += 200
                missing_slots = sum(1 for h in range(9, 19) if test_cov[h] == 0)
                combo_score -= missing_slots * 1000
                if combo_score > best_combo_score:
                    best_combo_score = combo_score
                    best_combo = combo

            if best_combo is None:
                raise ValueError(f"Impossibile assegnare turni Zetema per {DAY_LABELS[day_idx]}.")

            for op, shift in zip(zetema_workers, best_combo):
                schedule[op.name][day_idx] = shift
                preferred_shift.setdefault(op.name, shift)
                for h in shift_to_hours(shift):
                    coverage[h] += 1
                    if 9 <= h < 19:
                        z_coverage[h] += 1

        workers = [op for op in workers if op.group != "Zetema"]

        for op in workers:
            options = generate_shift_options(op, day_idx)
            if not options:
                raise ValueError(
                    f"Nessuna opzione di turno per '{op.name}' nel giorno {DAY_LABELS[day_idx]}."
                )

            random.shuffle(options)  # shuffle so equal-scored options vary across runs

            best_shift = None
            best_score = -10**9
            for option in options:
                hours = shift_to_hours(option)
                score = 0
                for h in hours:
                    unmet = max(0, demand_by_hour.get(h, [0] * 7)[day_idx] - coverage[h])
                    score += unmet * 10 + 1
                    if day_idx < 6 and coverage[h] == 0:
                        score += 180
                    if day_idx < 6 and op.group == "Zetema" and 9 <= h < 19 and z_coverage[h] == 0:
                        score += 260
                # Strong consistency bonus: same shift as the running preferred
                if preferred_shift.get(op.name) == option:
                    score += 300
                # Cross-week consistency: same shift as last week's dominant
                if prev_preferred and prev_preferred.get(op.name) == option:
                    score += 200
                if score > best_score:
                    best_score = score
                    best_shift = option

            assert best_shift is not None
            schedule[op.name][day_idx] = best_shift
            if op.name not in preferred_shift:
                preferred_shift[op.name] = best_shift
            for h in shift_to_hours(best_shift):
                coverage[h] += 1
                if op.group == "Zetema" and 9 <= h < 19:
                    z_coverage[h] += 1

        if day_idx < 6:
            all_day_workers = [op for op in operators if day_idx not in rest_days[op.name]]

            def recompute_day_coverage() -> Tuple[Dict[int, int], Dict[int, int]]:
                cov = {h: 0 for h in range(7, 20)}
                zcov = {h: 0 for h in range(9, 19)}
                for d_op in all_day_workers:
                    s = schedule[d_op.name][day_idx]
                    if not is_working_shift(s):
                        continue
                    for h in shift_to_hours(s):
                        cov[h] += 1
                        if d_op.group == "Zetema" and 9 <= h < 19:
                            zcov[h] += 1
                return cov, zcov

            coverage, z_coverage = recompute_day_coverage()
            missing_hours = [h for h in range(7, 20) if coverage[h] == 0]
            for missing_hour in missing_hours:
                fixed = False
                for op in all_day_workers:
                    old_shift = schedule[op.name][day_idx]
                    for option in generate_shift_options(op, day_idx):
                        if missing_hour not in shift_to_hours(option):
                            continue
                        schedule[op.name][day_idx] = option
                        test_cov, test_z = recompute_day_coverage()
                        if all(test_cov[h] >= 1 for h in range(7, 20)) and all(
                            test_z[h] >= 1 for h in range(9, 19)
                        ):
                            fixed = True
                            coverage, z_coverage = test_cov, test_z
                            break
                    if fixed:
                        break
                    schedule[op.name][day_idx] = old_shift
                if not fixed:
                    raise ValueError(
                        f"Impossibile garantire copertura minima per "
                        f"{DAY_LABELS[day_idx]} {missing_hour:02d}:00."
                    )

    return schedule


def validate_schedule(
    operators: List[Operator],
    schedule: Dict[str, List[str]],
    rest_days: Dict[str, Set[int]] | None = None,
) -> None:
    by_name = {op.name: op for op in operators}

    for op in operators:
        shifts = schedule.get(op.name)
        if not shifts or len(shifts) != 7:
            raise ValueError(f"Planning mancante o incompleto per '{op.name}'.")

        lav = sum(1 for s in shifts if is_working_shift(s))
        expected_lav = (7 - len(rest_days[op.name])) if rest_days else 5
        if lav != expected_lav:
            raise ValueError(
                f"'{op.name}': attesi {expected_lav} giorni lavorativi, trovati {lav}."
            )
        if op.group == "Non Zetema" and is_working_shift(shifts[6]):
            raise ValueError(f"'{op.name}' Non Zetema non puo' lavorare di domenica.")
        if op.group == "Zetema" and is_working_shift(shifts[6]):
            rip_lun_sab = sum(1 for s in shifts[:6] if not is_working_shift(s))
            # Per Zetema domenicale, rest_days contiene solo giorni lun-sab
            expected_rip_lun_sab = len(rest_days[op.name]) if rest_days else 2
            if rip_lun_sab != expected_rip_lun_sab:
                raise ValueError(f"'{op.name}' Zetema domenicale deve avere {expected_rip_lun_sab} RIP lun-sab.")

        for day_idx, shift in enumerate(shifts):
            if not is_working_shift(shift):
                continue
            start, end, pause = parse_shift(shift)
            if not (7 * 60 <= start < end <= 20 * 60):
                raise ValueError(f"Turno fuori fascia per '{op.name}' {DAY_LABELS[day_idx]}.")
            duration = end - start
            if op.daily_hours in (4, 6):
                if pause is not None or duration != op.daily_hours * 60:
                    raise ValueError(
                        f"Turno continuo non valido per '{op.name}' {DAY_LABELS[day_idx]}."
                    )
            else:
                if day_idx == 6:
                    raise ValueError(f"Turno 8h domenicale non ammesso per '{op.name}'.")
                if pause is None or duration != 9 * 60:
                    raise ValueError(f"Turno 8h non valido per '{op.name}' {DAY_LABELS[day_idx]}.")
                if start not in {7 * 60, 8 * 60, 9 * 60, 10 * 60, 11 * 60}:
                    raise ValueError(f"Inizio 8h non ammesso per '{op.name}'.")
                if pause - start not in {180, 240, 300}:
                    raise ValueError(f"Pausa 8h non ammessa per '{op.name}'.")

    for day_idx in range(7):
        cov = {h: 0 for h in range(7, 20)}
        zcov = {h: 0 for h in range(9, 19)}
        sunday_workers_local: List[Operator] = []

        for name, shifts in schedule.items():
            shift = shifts[day_idx]
            if not is_working_shift(shift):
                continue
            op = by_name[name]
            if day_idx == 6:
                sunday_workers_local.append(op)
                if op.group != "Zetema":
                    raise ValueError(f"Domenica puo' lavorare solo Zetema: '{name}'.")
            for h in shift_to_hours(shift):
                cov[h] += 1
                if op.group == "Zetema" and 9 <= h < 19:
                    zcov[h] += 1

        if day_idx < 6:
            for h in range(7, 20):
                if cov[h] < 1:
                    raise ValueError(
                        f"Copertura minima assente {DAY_LABELS[day_idx]} {h:02d}:00-{h+1:02d}:00."
                    )
            for h in range(9, 19):
                if zcov[h] < 1:
                    raise ValueError(
                        f"Manca Zetema {DAY_LABELS[day_idx]} {h:02d}:00-{h+1:02d}:00."
                    )
        else:
            for h in range(9, 19):
                if cov[h] < 1:
                    raise ValueError("Copertura domenica 09:00-19:00 non continua.")
            if len(sunday_workers_local) == 2:
                hours = sorted([w.daily_hours for w in sunday_workers_local])
                if hours != [4, 6]:
                    raise ValueError("Domenica con 2 operatori deve essere combinazione 6h + 4h.")
            elif len(sunday_workers_local) == 3:
                if any(w.daily_hours != 4 for w in sunday_workers_local):
                    raise ValueError("Domenica con 3 operatori ammesso solo 4h+4h+4h.")
            else:
                raise ValueError("Domenica numero operatori non ammesso.")


def compute_coverage_stats(
    operators: List[Operator],
    schedule: Dict[str, List[str]],
    demand_by_hour: Dict[int, List[int]],
) -> Tuple[Dict[int, List[int]], Dict[int, List[int]], Dict[int, List[int]], List[int]]:
    by_name = {op.name: op for op in operators}
    deployed = {h: [0] * 7 for h in range(7, 20)}
    z_deployed = {h: [0] * 7 for h in range(9, 19)}

    for day_idx in range(7):
        for name, shifts in schedule.items():
            shift = shifts[day_idx]
            if not is_working_shift(shift):
                continue
            op = by_name[name]
            for h in shift_to_hours(shift):
                if 7 <= h <= 19:
                    deployed[h][day_idx] += 1
                if op.group == "Zetema" and 9 <= h < 19:
                    z_deployed[h][day_idx] += 1

    delta = {h: [deployed[h][d] - demand_by_hour[h][d] for d in range(7)] for h in range(7, 20)}
    total_deficit = [sum(min(0, delta[h][d]) for h in range(7, 20)) for d in range(7)]
    return deployed, z_deployed, delta, total_deficit


def write_delta_section(
    ws,
    start_row: int,
    days: List[datetime],
    demand_by_hour: Dict[int, List[int]],
    num_operators: int,
) -> int:
    last_op_row = 3 + num_operators

    def dep(col: str, h: int) -> str:
        """Count operators whose shift covers hour h using pure text comparison.
        Shifts are zero-padded ("07", "16") so lexicographic order equals numeric order.
        No VALUE() needed — avoids array-eval errors in all Excel versions."""
        c = f"{col}$4:{col}${last_op_row}"
        hs = f"{h:02d}"
        return (
            f"SUMPRODUCT("
            f"(LEFT({c},2)<=\"{hs}\")"
            f"*(MID({c},7,2)>\"{hs}\")"
            f"*(IFERROR(IF(LEN({c})>11,MID({c},20,2),\"99\"),\"99\")<>\"{hs}\")"
            f")"
        )

    def zdep(col: str, h: int) -> str:
        """Same but restricted to Zetema group (column B)."""
        c = f"{col}$4:{col}${last_op_row}"
        b = f"$B$4:$B${last_op_row}"
        hs = f"{h:02d}"
        return (
            f"SUMPRODUCT("
            f"({b}=\"Zetema\")"
            f"*(LEFT({c},2)<=\"{hs}\")"
            f"*(MID({c},7,2)>\"{hs}\")"
            f"*(IFERROR(IF(LEN({c})>11,MID({c},20,2),\"99\"),\"99\")<>\"{hs}\")"
            f")"
        )

    red   = PatternFill("solid", fgColor="F8CBAD")
    green = PatternFill("solid", fgColor="E2F0D9")
    blue  = PatternFill("solid", fgColor="D9E2F3")
    center = Alignment(horizontal="center")

    ws.cell(row=start_row, column=1, value="DELTA FABBISOGNO vs DEPLOYED - Copertura Oraria").font = Font(
        bold=True, size=12, color="1F4E79"
    )
    ws.cell(row=start_row + 1, column=1, value="Risultato: Delta (Deployed − Fabbisogno)")

    headers = ["Fascia Oraria"] + [f"{DAY_LABELS[i]} {d.day}/{d.month}" for i, d in enumerate(days)]
    for c, value in enumerate(headers, start=1):
        cell = ws.cell(row=start_row + 2, column=c, value=value)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E79")
        cell.alignment = center

    row = start_row + 3
    delta_first = row
    for h in range(7, 20):
        ws.cell(row=row, column=1, value=f"{h:02d}:00-{h+1:02d}:00")
        for d in range(7):
            col = chr(ord("D") + d)
            fab = demand_by_hour[h][d]
            d_formula = dep(col, h)
            formula = f'=TEXT({d_formula}-{fab},"+0;-0;0")'
            cell = ws.cell(row=row, column=2 + d, value=formula)
            cell.alignment = center
        row += 1
    delta_last = row - 1

    # Colori condizionali: rosso se negativo, blu se positivo, verde se zero
    delta_range = f"B{delta_first}:H{delta_last}"
    ws.conditional_formatting.add(delta_range,
        FormulaRule(formula=[f'LEFT(B{delta_first},1)="-"'], fill=red))
    ws.conditional_formatting.add(delta_range,
        FormulaRule(formula=[f'LEFT(B{delta_first},1)="+"'], fill=blue))
    ws.conditional_formatting.add(delta_range,
        FormulaRule(formula=[f'B{delta_first}="0"'], fill=green))

    # Totale Deficit: somma MIN(0, deployed-fabbisogno) per ogni ora
    ws.cell(row=row, column=1, value="TOTALE DEFICIT").font = Font(bold=True)
    for d in range(7):
        col = chr(ord("D") + d)
        parts = [f"MIN(0,{dep(col,h)}-{demand_by_hour[h][d]})" for h in range(7, 20)]
        cell = ws.cell(row=row, column=2 + d, value="=" + "+".join(parts))
        cell.font = Font(bold=True)
        cell.alignment = center

    row += 3
    ws.cell(
        row=row, column=1, value="COPERTURA ZETEMA 09:00-19:00 (Lun-Sab)"
    ).font = Font(bold=True, size=12, color="1F4E79")
    row += 1
    for c, value in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=c, value=value)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E79")
        cell.alignment = center
    row += 1

    zet_first = row
    for h in range(9, 19):
        ws.cell(row=row, column=1, value=f"{h:02d}:00-{h+1:02d}:00")
        for d in range(7):
            col = chr(ord("D") + d)
            cell = ws.cell(row=row, column=2 + d, value=f"={zdep(col, h)}")
            cell.alignment = center
        row += 1
    zet_last = row - 1

    # CF Zetema: verde >=1, rosso <1 (Lun-Sab); domenica sempre blu
    ws.conditional_formatting.add(f"B{zet_first}:G{zet_last}",
        CellIsRule(operator="greaterThanOrEqual", formula=["1"], fill=green))
    ws.conditional_formatting.add(f"B{zet_first}:G{zet_last}",
        CellIsRule(operator="lessThan", formula=["1"], fill=red))
    ws.conditional_formatting.add(f"H{zet_first}:H{zet_last}",
        CellIsRule(operator="greaterThanOrEqual", formula=["0"], fill=blue))

    return row + 1


def _write_week_sheet(
    ws,
    operators: List[Operator],
    days: List[datetime],
    schedule: Dict[str, List[str]],
    demand_by_hour: Dict[int, List[int]],
    week_num: int,
) -> None:
    """Write a full week (grid + coverage tables) onto an already-created sheet."""
    title = (
        f"GRIGLIA TURNAZIONE - Settimana {week_num}: "
        f"{days[0].strftime('%d/%m/%Y')} → {days[-1].strftime('%d/%m/%Y')}"
    )
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=14, color="1F4E79")

    headers = (
        ["Operatore", "Gruppo", "Ore"]
        + [f"{DAY_LABELS[i]} {d.day}/{d.month}" for i, d in enumerate(days)]
        + ["Ore Sett."]
    )
    for col, value in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col, value=value)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E79")
        cell.alignment = Alignment(horizontal="center")

    for r, op in enumerate(operators, start=4):
        shifts = schedule[op.name]
        work_shifts = [s for s in shifts if is_working_shift(s)]
        dominant = max(set(work_shifts), key=work_shifts.count) if work_shifts else None
        weekly_hours = sum(op.daily_hours for s in shifts if is_working_shift(s))
        row_values = [op.name, op.group, op.daily_hours] + shifts + [weekly_hours]
        for c, value in enumerate(row_values, start=1):
            cell = ws.cell(row=r, column=c, value=value)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            if 4 <= c <= 10:
                if value == "RIP":
                    cell.fill = PatternFill("solid", fgColor="F2DCDB")
                    cell.font = Font(color="C00000", bold=True)
                elif not is_working_shift(value):
                    # Assenza con tipo (Ferie, Malattia, ecc.)
                    cell.fill = PatternFill("solid", fgColor="FFE0B2")
                    cell.font = Font(color="BF5000", bold=True)
                elif dominant is not None and value != dominant:
                    # Turno diverso da quello dominante della settimana
                    cell.fill = PatternFill("solid", fgColor="FFE699")
                    cell.font = Font(color="7F6000", bold=True)
                else:
                    cell.fill = PatternFill("solid", fgColor="D9E2F3")
                    cell.font = Font(color="1F4E79")
        ws.cell(row=r, column=1).alignment = Alignment(horizontal="left")
        if op.group == "Zetema":
            for c in range(1, 4):
                cell = ws.cell(row=r, column=c)
                cell.fill = PatternFill("solid", fgColor="E2EFDA")
                cell.font = Font(color="375623")

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 6
    for col in ["D", "E", "F", "G", "H", "I", "J"]:
        ws.column_dimensions[col].width = 20
    ws.column_dimensions["K"].width = 10

    write_delta_section(ws, 20, days, demand_by_hour, len(operators))


def write_report(
    operators: List[Operator],
    week_results: List[Tuple[List[datetime], Dict[str, List[str]], Dict[int, List[int]]]],
) -> Path:
    report_path = _HERE / "output" / "report_turnazione.txt"
    report_path.parent.mkdir(parents=True, exist_ok=True)
    lines: List[str] = []

    first_day = week_results[0][0][0]
    last_day = week_results[-1][0][-1]
    lines.append(
        f"Report turnazione: {first_day.strftime('%d/%m/%Y')} - {last_day.strftime('%d/%m/%Y')}"
    )
    lines.append(f"Numero settimane: {len(week_results)}")
    lines.append("")

    total_warnings = 0
    for week_idx, (days, schedule, demand_by_hour) in enumerate(week_results, start=1):
        deployed, z_deployed, delta, total_deficit = compute_coverage_stats(
            operators, schedule, demand_by_hour
        )
        lines.append(
            f"=== Settimana {week_idx}: "
            f"{days[0].strftime('%d/%m/%Y')} - {days[-1].strftime('%d/%m/%Y')} ==="
        )
        total_negative = sum(-v for v in total_deficit if v < 0)
        lines.append(f"Deficit totale ore-slot: {total_negative}")
        for d in range(7):
            lines.append(f"- {DAY_LABELS[d]} deficit: {total_deficit[d]}")

        lines.append("")
        lines.append("Warning copertura minima lun-sab:")
        for d in range(6):
            for h in range(7, 20):
                if deployed[h][d] < 1:
                    total_warnings += 1
                    lines.append(f"- {DAY_LABELS[d]} {h:02d}:00-{h+1:02d}:00 senza copertura minima")

        lines.append("")
        lines.append("Warning copertura Zetema lun-sab (09:00-19:00):")
        for d in range(6):
            for h in range(9, 19):
                if z_deployed[h][d] < 1:
                    total_warnings += 1
                    lines.append(f"- {DAY_LABELS[d]} {h:02d}:00-{h+1:02d}:00 senza Zetema")

        lines.append("")
        lines.append("Omogeneita' turni settimanale (soft):")
        for op in operators:
            work_shifts = [s for s in schedule[op.name] if is_working_shift(s)]
            distinct = len(set(work_shifts))
            lines.append(f"- {op.name}: {distinct} turno/i distinti")

        lines.append("")

    lines.append(f"Totale warning complessivi: {total_warnings}")
    report_path.write_text("\n".join(lines), encoding="utf-8")
    return report_path


def write_output(
    operators: List[Operator],
    week_results: List[Tuple[List[datetime], Dict[str, List[str]], Dict[int, List[int]]]],
) -> Tuple[Path, Path]:
    OUTPUT_FILE.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()

    # Replace the default empty sheet with the first week, add remaining weeks
    for week_idx, (days, schedule, demand_by_hour) in enumerate(week_results, start=1):
        if week_idx == 1:
            ws = wb.active
            ws.title = f"Sett {week_idx} ({days[0].strftime('%d.%m')})"
        else:
            ws = wb.create_sheet(title=f"Sett {week_idx} ({days[0].strftime('%d.%m')})")
        _write_week_sheet(ws, operators, days, schedule, demand_by_hour, week_idx)

    output_path = OUTPUT_FILE
    try:
        wb.save(output_path)
    except PermissionError:
        suffix = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = Path("output") / f"turnazione_generata_{suffix}.xlsx"
        wb.save(output_path)

    report_path = write_report(operators, week_results)
    return output_path, report_path


def main() -> None:
    operators, weeks, demand_schedule, absences = load_input(INPUT_FILE)
    zetema_ops = [op for op in operators if op.group == "Zetema"]

    # Rotation trackers across weeks
    # sunday_history[name]  = number of Sundays worked so far
    # work_history[name][day_of_week] = number of times that day was a working day
    sunday_history: Dict[str, int] = {op.name: 0 for op in operators}
    work_history: Dict[str, Dict[int, int]] = {op.name: {} for op in operators}
    prev_preferred: Dict[str, str] | None = None

    week_results: List[Tuple[List[datetime], Dict[str, List[str]], Dict[int, List[int]]]] = []

    for week_num, week_days in enumerate(weeks, start=1):
        demand_by_hour = get_demand_for_week(demand_schedule, week_days[0])
        week_absences = get_week_absences(absences, week_days)

        best_schedule = None
        best_rest_days = None
        best_sunday_workers = None
        best_deficit_score: float = float("inf")
        attempts_done = 0

        for attempt in range(MAX_SCHEDULE_RETRIES):
            attempts_done = attempt + 1
            sunday_workers = pick_sunday_workers(zetema_ops, week_absences, sunday_history)
            rest_days = generate_rest_days(
                operators, demand_by_hour, week_absences, sunday_workers, work_history
            )
            schedule = build_schedule(
                operators, demand_by_hour, rest_days, sunday_workers, prev_preferred
            )

            _, _, _, total_deficit = compute_coverage_stats(
                operators, schedule, demand_by_hour
            )
            deficit_score = sum(-d for d in total_deficit if d < 0)

            if deficit_score < best_deficit_score:
                best_deficit_score = deficit_score
                best_schedule = {name: list(shifts) for name, shifts in schedule.items()}
                best_rest_days = {name: set(days) for name, days in rest_days.items()}
                best_sunday_workers = set(sunday_workers)

            if all(d >= DEFICIT_TOLERANCE for d in total_deficit):
                break

        schedule = best_schedule
        rest_days = best_rest_days
        sunday_workers = best_sunday_workers

        if best_deficit_score > 0 and attempts_done > 1:
            print(
                f"  [AVVISO] Sett. {week_num}: miglior risultato dopo {attempts_done} tentativi"
                f" — deficit residuo {best_deficit_score} persona-ore."
            )

        # Replace "RIP" on absence days with the actual tipo label (Ferie, Malattia, ecc.)
        for op in operators:
            for day_idx, tipo in week_absences.get(op.name, {}).items():
                if schedule[op.name][day_idx] == "RIP":
                    schedule[op.name][day_idx] = tipo

        validate_schedule(operators, schedule, rest_days)

        # Update rotation trackers for next week
        for op in operators:
            if is_working_shift(schedule[op.name][6]):
                sunday_history[op.name] += 1
            for day_idx, shift in enumerate(schedule[op.name]):
                if is_working_shift(shift):
                    work_history[op.name][day_idx] = work_history[op.name].get(day_idx, 0) + 1

        # Remember this week's preferred shifts for variety scoring next week
        prev_preferred = {}
        for op in operators:
            work_shifts = [s for s in schedule[op.name] if is_working_shift(s)]
            if work_shifts:
                # Most common shift this week = "preferred" to vary next week
                prev_preferred[op.name] = max(set(work_shifts), key=work_shifts.count)

        week_results.append((week_days, schedule, demand_by_hour))
        print(
            f"Settimana {week_num} ({week_days[0].strftime('%d/%m')}–"
            f"{week_days[-1].strftime('%d/%m')}) generata."
        )

    output_path, report_path = write_output(operators, week_results)
    print(f"\nOutput generato: {output_path}")
    print(f"Report generato: {report_path}")


if __name__ == "__main__":
    main()
